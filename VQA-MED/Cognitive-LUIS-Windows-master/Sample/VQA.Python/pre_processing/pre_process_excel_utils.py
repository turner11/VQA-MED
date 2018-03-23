import argparse
import openpyxl
import shutil
import os
import re
import pandas as pd

from parsers.utils import suppress_func_stdout, has_word
from pre_processing.known_find_and_replace_items import find_and_replace_collection, locations, diagnosis, \
    imaging_devices
from parsers.VQA18 import Vqa18_from_excel, Vqa18_from_raw_csv, Vqa18Base, TOKENIZED_COL_PREFIX
from vqa_logger import logger


class ExcelPreProcessor(object):
    """"""


    def __init__(self):
        """"""
        super().__init__()

    @classmethod
    def _find_and_replace(cls, path, find_and_replace_data, columns_filter=None):
        columns_filter = columns_filter or (lambda clm_name: True)

        excel_file = openpyxl.load_workbook(path)
        for sheet_name in excel_file.sheetnames:
            curr_sheet = excel_file.get_sheet_by_name(sheet_name)

            cels_in_first_row = list(curr_sheet.iter_rows())[0]
            cols = [(c.value, idx) for c, idx in zip(cels_in_first_row, range(len(cels_in_first_row)))]
            cols_idxs = [c[1] for c in cols if columns_filter(c[0])]

            current_row = 1
            for row in curr_sheet.iter_rows():
                for col_idx in range(len(row)):
                    if col_idx not in cols_idxs:
                        continue
                    seed_one_col = col_idx + 1
                    curr_cell = curr_sheet.cell(row=current_row, column=seed_one_col)
                    val = str(curr_cell.value)
                    if not val or val == 'None':
                        continue
                    new_val = val
                    for tpl in find_and_replace_data:
                        pattern = re.compile(tpl.orig, re.IGNORECASE)
                        new_val = pattern.sub(tpl.sub, new_val)
                        # new_val = new_val.replace(tpl.orig, tpl.sub)
                    curr_cell.value = new_val.strip()
                current_row += 1

        excel_file.save(path)

    @classmethod
    def process_excel(cls, path):
        find_and_replace_data = find_and_replace_collection
        col_filter = lambda col_name: TOKENIZED_COL_PREFIX in (col_name or '')
        cls._find_and_replace(path, find_and_replace_data, columns_filter=col_filter)

        df = pd.read_excel(path)
        cls.add_imaging_columns(df)
        cls.add_diagnostics_columns(df)
        cls.add_locations_columns(df)

        writer = pd.ExcelWriter(path)
        df.to_excel(writer, 'vqa_data')
        writer.save()

        # df = pd.DataFrame({'Data': [10, 20, 30, 20, 15, 30, 45]})

        # # Create a Pandas Excel writer using XlsxWriter as the engine.
        # writer = pd.ExcelWriter('pandas_conditional.xlsx', engine='xlsxwriter')
        #
        # # Convert the dataframe to an XlsxWriter Excel object.
        # df.to_excel(writer, sheet_name='Sheet1')
        #
        # # Get the xlsxwriter workbook and worksheet objects.
        # workbook = writer.book
        # worksheet = writer.sheets['Sheet1']
        #
        # # Apply a conditional format to the cell range.
        # worksheet.conditional_format('B2:B8', {'type': '3_color_scale'})
        #
        # # Close the Pandas Excel writer and output the Excel file.
        # writer.save()

    @classmethod
    def add_locations_columns(cls, df):
        cls.add_columns_by_search(df
                                  , indicator_words=locations
                                  , search_columns=[Vqa18Base.COL_TOK_Q, Vqa18Base.COL_TOK_A, Vqa18Base.COL_QUESTION, Vqa18Base.COL_ANSWER])

    @classmethod
    def add_diagnostics_columns(cls, df):
        cls.add_columns_by_search(df
                                  , indicator_words=diagnosis
                                  , search_columns=[Vqa18Base.COL_TOK_Q, Vqa18Base.COL_TOK_A, Vqa18Base.COL_QUESTION, Vqa18Base.COL_ANSWER])

    @classmethod
    def add_imaging_columns(cls, df):
        cls.add_columns_by_search(df
                                  , indicator_words=imaging_devices
                                  , search_columns=[Vqa18Base.COL_TOK_Q, Vqa18Base.COL_TOK_A])

    @classmethod
    def add_columns_by_search(cls, df, indicator_words, search_columns):
        for word in indicator_words:
            res = None
            for col in search_columns:
                curr_res = df[col].apply(lambda s: has_word(word,s))
                if res is None:
                    res = curr_res
                res = res | curr_res
            if any(res):
                df[word] = res
            else:
                logger.warn("found no matching for '{0}'".format(word))




class DataPreProcessor(object):
    """"""

    def __init__(self):
        """"""
        super().__init__()

    @classmethod
    def preprocessed_data__to_processed_excel(cls, raw_data_path, output_excel_path):
        is_csv = raw_data_path.lower().endswith('csv')
        is_excel = any(raw_data_path.lower().endswith(ext) for ext in ['xlsx', 'xls'])
        vqa_instance = None
        if is_csv:
            vqa_instance = cls.raw_csv_to_processed_excel(raw_data_path, output_excel_path)
        elif is_excel:
            vqa_instance = cls._excel_to_processed(raw_data_path, output_excel_path)
        else:
            raise NotImplementedError("Could not pre process file: {0}".format(raw_data_path))
        return vqa_instance
    @classmethod
    def raw_csv_to_processed_excel(cls, raw_csv_path, output_excel_path):
        """
        Getting a processed excel from raw csv.
        :param raw_csv_path: path of raw csv
        :param output_excel_path: path of out put csv
        :return: a data parser contingin processed data
        """
        inst = Vqa18_from_raw_csv(raw_csv_path)
        data = inst.data
        cols_to_tokenize = [
            (Vqa18Base.COL_QUESTION, Vqa18Base.COL_TOK_Q),
            (Vqa18Base.COL_ANSWER, Vqa18Base.COL_TOK_A)
        ]
        for col, new_tok_col in cols_to_tokenize:
            cls.add_tokenized_column(data, col, new_tok_col)

        def append_id(filename, suffix):
            name, ext = os.path.splitext(filename)
            return "{name}_{uid}{ext}".format(name=name, uid=suffix, ext=ext)

        intermediate_excel_path = append_id(output_excel_path, "intermediate")
        cls._dump_to_excel(data, intermediate_excel_path)
        logger.debug("Original CSV data:\n\t{0}\nProcessed Data\n\t{1}".format(raw_csv_path, output_excel_path))
        return cls._excel_to_processed(intermediate_excel_path, output_excel_path)

    @classmethod
    def _excel_to_processed(cls, excel_path, output_processed_excel_path):
        shutil.copy(excel_path, output_processed_excel_path)
        ExcelPreProcessor.process_excel(output_processed_excel_path)
        logger.debug("Original data:\n\t{0}\nProcessed Data\n\t{1}".format(excel_path,output_processed_excel_path))
        return Vqa18_from_excel(output_processed_excel_path)

    @classmethod
    def _dump_to_excel(cls, df, path):
        from pandas import ExcelWriter
        writer = ExcelWriter(path)
        df.to_excel(writer, 'vqa_data')
        writer.save()

    @classmethod
    @suppress_func_stdout
    def add_tokenized_column(cls, df, source_col, dest_col):
        from parsers.token_utils import Preprocessor
        df[dest_col] = df[source_col].apply(lambda s: " ".join(Preprocessor.tokeniz(s)))


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='')
    parser.add_argument('-p', dest='path', help='path to excel', default='')
    parser.add_argument('-d', dest='destination', help='destination of processed file', default='')

    args = parser.parse_args()
    dbg_file_xls = 'C:\\Users\\avitu\\Documents\\GitHub\\VQA-MED\\VQA-MED\\Cognitive-LUIS-Windows-master\\Sample\\VQA.Python\\dumped_data\\vqa_data.xlsx'
    dbg_file_csv = 'C:\\Users\\Public\\Documents\\Data\\2018\\VQAMed2018Train\\VQAMed2018Train-QA.csv'

    args.path = args.path or dbg_file_csv

    args.destination = args.destination or os.path.splitext(args.path)[0] + '_post_pre_process.xlsx'
    DataPreProcessor.preprocessed_data__to_processed_excel(args.path, args.destination)
